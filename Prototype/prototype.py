import openpyxl
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook('excel-sheets/Lab-output.xlsx')

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have a row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method.
max_column = sheet_obj.max_column
labels = ['Id', 'Name: ', 'protein_sample 1: ', 'protein_sample 2: ', 'protein_sample 3: ', 'protein_sample 4: ', 'Sex: ', 'Age: ']

# Create a new PDF file named output.pdf
pdf_file = canvas.Canvas('output.pdf', pagesize=letter)

# Set x-coordinate value for protein labels and values
protein_x = 400

# Set y-coordinate values for Sex and Age labels
sex_y = 640
age_y = 600

for row in sheet_obj.iter_rows(min_row=2):
    values = {}
    for i, cell in enumerate(row):
        if i >= len(labels):
            break
        values[f'value_{i}'] = cell.value

        # Display labels and values for id, name, sex, and age
        if i in [0, 1]:
            pdf_file.drawString(72, 720-40*i, f"{labels[i]} {cell.value}")
        elif i == 6:
            pdf_file.drawString(72, sex_y, f"{labels[i]} {cell.value}")
        elif i == 7:
            pdf_file.drawString(72, age_y, f"{labels[i]} {cell.value}")
        # Display protein labels and values on the right side of the screen
        elif i in [2, 3, 4, 5]:
            pdf_file.drawString(protein_x, 720-40*(i-1), f"{labels[i]} {cell.value}")

    value_2 = values['value_2'] * 2
    value_3 = values['value_3'] * 0.5
    value_4 = values['value_4'] * 1
    value_5 = values['value_5'] * 0.5
    value_7 = values['value_7'] * 0.5
    all = value_2 + value_3 + value_4 + value_5 + value_7
    pdf_file.drawString(72, 200, f"Total Score: {all}")
    if all < 30:
        result = 'Low'
    elif all >= 30 and all < 40:
        result = 'Medium'
    elif all >= 40:
        result = 'High'
    pdf_file.drawString(72, 180, f"Risk Factor: {result}")
    
    sheet_obj.cell(row=row[0].row, column=sheet_obj.max_column, value=result)
    sheet_obj.cell(row=row[0].row, column=sheet_obj.max_column - 1, value=all)

    pdf_file.showPage()

# Print value of cell object
# using the value attribute
wb_obj.save('excel-sheets/Lab-output.xlsx')

pdf_file.save()

print(max_column)